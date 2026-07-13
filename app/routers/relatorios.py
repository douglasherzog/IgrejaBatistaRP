import csv
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Request, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import extract, func
from datetime import date
from collections import defaultdict
from app.database import get_db
from app.models import Lancamento, TipoLancamento, Membro, Evento, InscricaoEvento, PedidoOracao, Admin
from app.auth import get_admin_atual
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

router = APIRouter()
templates = Jinja2Templates(directory="templates")

MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
MESES_FULL = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
              "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


@router.get("/admin/relatorios/caixa", response_class=HTMLResponse)
def relatorio_caixa(
    request: Request,
    ano: int = Query(default=None),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    hoje = date.today()
    ano = ano or hoje.year

    lancamentos_ano = db.query(Lancamento).filter(
        extract("year", Lancamento.data) == ano
    ).all()

    receitas_mes = [0.0] * 12
    despesas_mes = [0.0] * 12
    cat_receitas = defaultdict(float)
    cat_despesas = defaultdict(float)

    for l in lancamentos_ano:
        m = l.data.month - 1
        if l.tipo == TipoLancamento.receita:
            receitas_mes[m] += float(l.valor)
            cat_receitas[l.categoria] += float(l.valor)
        else:
            despesas_mes[m] += float(l.valor)
            cat_despesas[l.categoria] += float(l.valor)

    total_receitas = sum(receitas_mes)
    total_despesas = sum(despesas_mes)
    saldo_ano = total_receitas - total_despesas

    return templates.TemplateResponse("admin/relatorios/caixa.html", {
        "request": request,
        "ano": ano,
        "receitas_mes": receitas_mes,
        "despesas_mes": despesas_mes,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "saldo_ano": saldo_ano,
        "cat_receitas": dict(cat_receitas),
        "cat_despesas": dict(cat_despesas),
        "meses": MESES,
    })


@router.get("/admin/relatorios/caixa/csv")
def relatorio_caixa_csv(
    ano: int = Query(default=None),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    hoje = date.today()
    ano = ano or hoje.year

    lancamentos = db.query(Lancamento).filter(
        extract("year", Lancamento.data) == ano
    ).order_by(Lancamento.data).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Data", "Tipo", "Categoria", "Descricao", "Valor", "Dizimista"])
    for l in lancamentos:
        dizimista = ""
        if l.categoria == "Dízimo" and l.membro:
            dizimista = f"{l.membro.nome} {l.membro.sobrenome}"
        writer.writerow([
            l.data.strftime("%d/%m/%Y"),
            l.tipo.value,
            l.categoria,
            l.descricao,
            f"R$ {float(l.valor):.2f}",
            dizimista,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=caixa_{ano}.csv"}
    )


@router.get("/admin/relatorios/membros", response_class=HTMLResponse)
def relatorio_membros(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    hoje = date.today()
    total = db.query(Membro).count()
    ativos = db.query(Membro).filter(Membro.ativo == True).count()
    inativos = total - ativos

    batismos_ano = db.query(Membro).filter(
        extract("year", Membro.data_batismo) == hoje.year
    ).count()

    novos_ano = db.query(Membro).filter(
        extract("year", Membro.criado_em) == hoje.year
    ).count()

    batismos_por_mes = [0] * 12
    for m in db.query(Membro).filter(Membro.data_batismo.isnot(None)).all():
        if m.data_batismo.year == hoje.year:
            batismos_por_mes[m.data_batismo.month - 1] += 1

    novos_por_mes = [0] * 12
    for m in db.query(Membro).all():
        if m.criado_em and m.criado_em.year == hoje.year:
            novos_por_mes[m.criado_em.month - 1] += 1

    aniversariantes = db.query(Membro).filter(
        Membro.ativo == True,
        Membro.data_nascimento.isnot(None),
        extract("month", Membro.data_nascimento) == hoje.month
    ).order_by(extract("day", Membro.data_nascimento)).all()

    return templates.TemplateResponse("admin/relatorios/membros.html", {
        "request": request,
        "total": total,
        "ativos": ativos,
        "inativos": inativos,
        "batismos_ano": batismos_ano,
        "novos_ano": novos_ano,
        "batismos_por_mes": batismos_por_mes,
        "novos_por_mes": novos_por_mes,
        "aniversariantes": aniversariantes,
        "meses": MESES,
        "mes_atual": MESES_FULL[hoje.month - 1],
    })


@router.get("/admin/relatorios/membros/csv")
def relatorio_membros_csv(
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    membros = db.query(Membro).order_by(Membro.nome).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nome", "Sobrenome", "Nascimento", "Batismo", "Telefone", "Email", "Cidade", "Status"])
    for m in membros:
        writer.writerow([
            m.nome,
            m.sobrenome,
            m.data_nascimento.strftime("%d/%m/%Y") if m.data_nascimento else "",
            m.data_batismo.strftime("%d/%m/%Y") if m.data_batismo else "",
            m.telefone or "",
            m.email or "",
            m.cidade or "",
            "Ativo" if m.ativo else "Inativo",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=membros.csv"}
    )


@router.get("/admin/relatorios/geral", response_class=HTMLResponse)
def relatorio_geral(
    request: Request,
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    hoje = date.today()

    total_membros = db.query(Membro).count()
    novos_mes = db.query(Membro).filter(
        extract("month", Membro.criado_em) == hoje.month,
        extract("year", Membro.criado_em) == hoje.year
    ).count()

    lanc_ano = db.query(Lancamento).filter(extract("year", Lancamento.data) == hoje.year).all()
    rec_mes = [0.0] * 12
    desp_mes = [0.0] * 12
    for l in lanc_ano:
        idx = l.data.month - 1
        if l.tipo == TipoLancamento.receita:
            rec_mes[idx] += float(l.valor)
        else:
            desp_mes[idx] += float(l.valor)

    eventos_futuros = db.query(Evento).filter(Evento.data >= hoje, Evento.ativo == True).count()
    total_inscricoes = db.query(InscricaoEvento).count()
    pedidos_novos = db.query(PedidoOracao).filter(PedidoOracao.status == "novo").count()
    total_pedidos = db.query(PedidoOracao).count()

    return templates.TemplateResponse("admin/relatorios/geral.html", {
        "request": request,
        "total_membros": total_membros,
        "novos_mes": novos_mes,
        "rec_mes": rec_mes,
        "desp_mes": desp_mes,
        "eventos_futuros": eventos_futuros,
        "total_inscricoes": total_inscricoes,
        "pedidos_novos": pedidos_novos,
        "total_pedidos": total_pedidos,
        "meses": MESES,
        "ano": hoje.year,
    })


# ===== EXPORTAÇÃO EXCEL =====

@router.get("/admin/relatorios/caixa/excel")
def exportar_caixa_excel(
    ano: int = Query(...),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    """Exporta relatório de caixa em Excel"""
    
    # Buscar dados do relatório
    lancamentos = db.query(Lancamento).filter(
        extract('year', Lancamento.data) == ano
    ).order_by(Lancamento.data).all()
    
    # Criar workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Caixa {ano}"
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    headers = ['Data', 'Tipo', 'Categoria', 'Descrição', 'Valor', 'Membro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados
    for row, lanc in enumerate(lancamentos, 2):
        ws.cell(row=row, column=1, value=lanc.data.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value='Receita' if lanc.tipo == TipoLancamento.receita else 'Despesa')
        ws.cell(row=row, column=3, value=lanc.categoria)
        ws.cell(row=row, column=4, value=lanc.descricao)
        ws.cell(row=row, column=5, value=float(lanc.valor))
        ws.cell(row=row, column=6, value=f"{lanc.membro.nome} {lanc.membro.sobrenome}" if lanc.membro else '-')
    
    # Ajustar colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em memória
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"relatorio_caixa_{ano}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/admin/relatorios/membros/excel")
def exportar_membros_excel(
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    """Exporta relatório de membros em Excel"""
    
    # Buscar dados
    membros = db.query(Membro).order_by(Membro.nome).all()
    
    # Criar workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membros"
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    headers = ['Nome', 'Sobrenome', 'Telefone', 'Email', 'Data Nascimento', 'Data Batismo', 'Cidade', 'Estado', 'Status', 'Data Cadastro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados
    for row, membro in enumerate(membros, 2):
        ws.cell(row=row, column=1, value=membro.nome)
        ws.cell(row=row, column=2, value=membro.sobrenome)
        ws.cell(row=row, column=3, value=membro.telefone or '')
        ws.cell(row=row, column=4, value=membro.email or '')
        ws.cell(row=row, column=5, value=membro.data_nascimento.strftime('%d/%m/%Y') if membro.data_nascimento else '')
        ws.cell(row=row, column=6, value=membro.data_batismo.strftime('%d/%m/%Y') if membro.data_batismo else '')
        ws.cell(row=row, column=7, value=membro.cidade or '')
        ws.cell(row=row, column=8, value=membro.estado or '')
        ws.cell(row=row, column=9, value='Ativo' if membro.ativo else 'Inativo')
        ws.cell(row=row, column=10, value=membro.data_cadastro.strftime('%d/%m/%Y') if membro.data_cadastro else '')
    
    # Ajustar colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em memória
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = "relatorio_membros.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ===== EXPORTAÇÃO PDF =====

@router.get("/admin/relatorios/caixa/pdf")
def exportar_caixa_pdf(
    ano: int = Query(...),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    """Exporta relatório de caixa em PDF"""
    
    # Buscar dados
    lancamentos = db.query(Lancamento).filter(
        extract('year', Lancamento.data) == ano
    ).order_by(Lancamento.data).all()
    
    # Criar buffer PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # centro
    )
    
    # Conteúdo
    story = []
    
    # Título
    story.append(Paragraph(f"Relatório de Caixa - {ano}", title_style))
    story.append(Spacer(1, 12))
    
    # Dados da tabela
    data = [['Data', 'Tipo', 'Categoria', 'Descrição', 'Valor']]
    
    for lanc in lancamentos:
        data.append([
            lanc.data.strftime('%d/%m/%Y'),
            'Receita' if lanc.tipo == TipoLancamento.receita else 'Despesa',
            lanc.categoria,
            lanc.descricao[:30] + '...' if len(lanc.descricao) > 30 else lanc.descricao,
            f"R$ {lanc.valor:.2f}"
        ])
    
    # Criar tabela
    table = Table(data, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 1*inch])
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = f"relatorio_caixa_{ano}.pdf"
    
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type='application/pdf',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/admin/relatorios/membros/pdf")
def exportar_membros_pdf(
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_admin_atual)
):
    """Exporta relatório de membros em PDF"""
    
    # Buscar dados
    membros = db.query(Membro).order_by(Membro.nome).all()
    
    # Criar buffer PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, landscape=True)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # centro
    )
    
    # Conteúdo
    story = []
    
    # Título
    story.append(Paragraph("Relatório de Membros", title_style))
    story.append(Spacer(1, 12))
    
    # Dados da tabela
    data = [['Nome', 'Telefone', 'Cidade', 'Status']]
    
    for membro in membros:
        data.append([
            f"{membro.nome} {membro.sobrenome}",
            membro.telefone or '-',
            membro.cidade or '-',
            'Ativo' if membro.ativo else 'Inativo'
        ])
    
    # Criar tabela
    table = Table(data, colWidths=[3*inch, 2*inch, 2*inch, 1*inch])
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    
    filename = "relatorio_membros.pdf"
    
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type='application/pdf',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
